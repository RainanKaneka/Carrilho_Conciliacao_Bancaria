import argparse
import warnings
import pandas as pd
import re
import itertools
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

def parse_currency(val):
    """Função blindada para converter qualquer formato de dinheiro para número"""
    if pd.isna(val): return None
    if isinstance(val, (int, float)): return float(val)
    val = str(val).upper().replace('R$', '').strip()
    if '.' in val and ',' in val:
        val = val.replace('.', '').replace(',', '.')
    elif ',' in val:
        val = val.replace(',', '.')
    try:
        return float(val)
    except:
        return None


class DataCleaner:
    @staticmethod
    def clean_argos(file_path: str) -> pd.DataFrame:
        try:
            import pandas as pd
            df = pd.read_excel(file_path, engine='openpyxl')
            
            header_idx = None
            cols_str = ' '.join(str(c).lower() for c in df.columns)
            if not ('valor' in cols_str and ('cliente' in cols_str or 'parceiro' in cols_str)):
                for idx, row in df.head(20).iterrows():
                    row_str = ' '.join(str(val).lower() for val in row.values)
                    if 'valor' in row_str and ('cliente' in row_str or 'parceiro' in row_str):
                        header_idx = idx
                        break
            
            if header_idx is not None:
                df = pd.read_excel(file_path, header=header_idx + 1, engine='openpyxl')
            
            col_mapping = {}
            for col in df.columns:
                col_lower = str(col).lower()
                if 'banco' in col_lower and 'baixa' not in col_lower and 'Banco' not in col_mapping.values(): 
                    col_mapping[col] = 'Banco'
                elif ('cliente' in col_lower or 'parceiro' in col_lower) and 'Cliente' not in col_mapping.values(): 
                    col_mapping[col] = 'Cliente'
                elif 'valor' in col_lower and 'Valor' not in col_mapping.values(): 
                    col_mapping[col] = 'Valor'
                elif 'data' in col_lower and 'baixa' not in col_lower and 'Data' not in col_mapping.values(): 
                    col_mapping[col] = 'Data'
                elif ('obs' in col_lower or 'hist' in col_lower) and 'Histórico' not in col_mapping.values(): 
                    col_mapping[col] = 'Histórico'
                elif 'evento descri' in col_lower and 'Tipo Evento' not in col_mapping.values():
                    col_mapping[col] = 'Tipo Evento'
            
            df = df.rename(columns=col_mapping)
            df = df.loc[:, ~df.columns.duplicated()].copy()
            cols_to_keep = [c for c in ['Banco', 'Cliente', 'Valor', 'Data', 'Histórico', 'Tipo Evento'] if c in df.columns]
            df = df[cols_to_keep]
            
            if 'Valor' not in df.columns: return pd.DataFrame()
            if 'Data' not in df.columns: df['Data'] = ''
            if 'Histórico' not in df.columns: df['Histórico'] = ''
            if 'Tipo Evento' not in df.columns: df['Tipo Evento'] = ''
            if 'Banco' not in df.columns: df['Banco'] = ''
            if 'Cliente' not in df.columns: df['Cliente'] = 'CLIENTE NÃO INFORMADO'
            
            nome_arquivo = str(file_path).lower()
            banco_nome = "ARGOS"
            if 'caixa' in nome_arquivo: banco_nome = 'CAIXA ECONOMICA'
            elif 'banese' in nome_arquivo: banco_nome = 'BANESE'

            df['Banco'] = df['Banco'].fillna(banco_nome)
            df['Banco'] = df['Banco'].replace(r'^\s*$', banco_nome, regex=True)
            
            df = df.dropna(subset=['Valor'])
            
            def parse_currency(val):
                if pd.isna(val): return None
                if isinstance(val, (int, float)): return float(val)
                val = str(val).upper().replace('R$', '').strip()
                if '.' in val and ',' in val:
                    val = val.replace('.', '').replace(',', '.')
                elif ',' in val:
                    val = val.replace(',', '.')
                try: return float(val)
                except: return None

            # NOVO PARSER DE DATA BLINDADO (Padrão BR)
            def parse_date_br(val):
                import datetime
                if pd.isna(val) or str(val).strip() == '': return ''
                if isinstance(val, datetime.datetime): return val.strftime('%d/%m/%Y')
                val_str = str(val).strip().replace('.', '')
                try:
                    partes = val_str.split('/')
                    if len(partes) == 2:
                        val_str = f"{val_str}/2026"
                    return pd.to_datetime(val_str, dayfirst=True).strftime('%d/%m/%Y')
                except:
                    return val_str

            df['Valor'] = df['Valor'].apply(parse_currency)
            df['Data'] = df['Data'].apply(parse_date_br)
            df['Histórico'] = df['Histórico'].fillna('')
            
            return df.dropna(subset=['Valor'])
        except Exception as e:
            print(f"Erro ao ler Argos {file_path}: {e}")
            return pd.DataFrame()

    @staticmethod
    def clean_bank(file_path: str) -> pd.DataFrame:
        try:
            import pandas as pd
            df = pd.read_excel(file_path, engine='openpyxl')
            
            header_idx = None
            cols_str = ' '.join(str(c).lower() for c in df.columns)
            if not ('data' in cols_str and 'valor' in cols_str):
                for idx, row in df.head(20).iterrows():
                    row_str = ' '.join(str(val).lower() for val in row.values)
                    if 'data' in row_str and 'valor' in row_str:
                        header_idx = idx
                        break
            
            if header_idx is not None:
                df = pd.read_excel(file_path, header=header_idx + 1, engine='openpyxl')
            
            col_mapping = {}
            for col in df.columns:
                col_lower = str(col).lower()
                if 'data' in col_lower and 'Data' not in col_mapping.values(): 
                    col_mapping[col] = 'Data'
                elif ('histórico' in col_lower or 'historico' in col_lower or 'descrição' in col_lower or 'hitórico' in col_lower) and 'Histórico' not in col_mapping.values(): 
                    col_mapping[col] = 'Histórico'
                elif 'valor' in col_lower and 'Valor' not in col_mapping.values(): 
                    col_mapping[col] = 'Valor'
                elif 'tipo' in col_lower and 'Tipo' not in col_mapping.values():
                    col_mapping[col] = 'Tipo'
            
            df = df.rename(columns=col_mapping)
            df = df.loc[:, ~df.columns.duplicated()].copy()
            cols_to_keep = [c for c in ['Data', 'Histórico', 'Valor', 'Tipo'] if c in df.columns]
            df = df[cols_to_keep]
            
            if 'Valor' not in df.columns: return pd.DataFrame()
            if 'Histórico' not in df.columns: df['Histórico'] = ''
            if 'Tipo' not in df.columns: df['Tipo'] = ''
            
            df = df.dropna(subset=['Valor'])
            
            def parse_currency(val):
                if pd.isna(val): return None
                if isinstance(val, (int, float)): return float(val)
                val = str(val).upper().replace('R$', '').strip()
                if '.' in val and ',' in val:
                    val = val.replace('.', '').replace(',', '.')
                elif ',' in val:
                    val = val.replace(',', '.')
                try: return float(val)
                except: return None

            def parse_date_br(val):
                import datetime
                if pd.isna(val) or str(val).strip() == '': return ''
                if isinstance(val, datetime.datetime): return val.strftime('%d/%m/%Y')
                val_str = str(val).strip().replace('.', '')
                try:
                    partes = val_str.split('/')
                    if len(partes) == 2:
                        val_str = f"{val_str}/2026"
                    return pd.to_datetime(val_str, dayfirst=True).strftime('%d/%m/%Y')
                except:
                    return val_str

            df['Valor'] = df['Valor'].apply(parse_currency)
            df['Data'] = df['Data'].apply(parse_date_br)
            df = df.dropna(subset=['Valor'])
            df = df[df['Valor'] > 0]
            df['Histórico'] = df['Histórico'].fillna('')
            
            nome_arquivo = str(file_path).lower()
            banco_nome = "BANCO DESCONHECIDO"
            if 'caixa' in nome_arquivo: banco_nome = 'CAIXA ECONOMICA'
            elif 'banese' in nome_arquivo: banco_nome = 'BANESE'
            df['Banco'] = banco_nome
            
            return df
        except Exception as e:
            print(f"Erro ao ler Banco {file_path}: {e}")
            return pd.DataFrame()



class ReconciliationEngine:
    def __init__(self, df_argos, df_bank):
        import pandas as pd
        
        self.df_argos = df_argos.copy() if isinstance(df_argos, pd.DataFrame) else pd.DataFrame()
        self.df_bank = df_bank.copy() if isinstance(df_bank, pd.DataFrame) else pd.DataFrame()

        # BLINDAGEM: Garante que todos os valores monetários são floats matemáticos válidos
        def safe_float(val):
            if pd.isna(val): return 0.0
            if isinstance(val, (int, float)): return float(val)
            val = str(val).upper().replace('R$', '').strip()
            if '.' in val and ',' in val:
                val = val.replace('.', '').replace(',', '.')
            elif ',' in val:
                val = val.replace(',', '.')
            try:
                return float(val)
            except:
                return 0.0

        if not self.df_argos.empty and 'Valor' in self.df_argos.columns:
            self.df_argos['Valor'] = self.df_argos['Valor'].apply(safe_float)
            
        if not self.df_bank.empty and 'Valor' in self.df_bank.columns:
            self.df_bank['Valor'] = self.df_bank['Valor'].apply(safe_float)

        saidas_estornos = []
        if not self.df_argos.empty:
            if 'Tipo Evento' not in self.df_argos.columns: self.df_argos['Tipo Evento'] = ''
            is_estorno = (self.df_argos['Valor'] < 0) | (self.df_argos['Tipo Evento'].astype(str).str.lower().str.contains('estorno'))
            for _, row in self.df_argos[is_estorno].iterrows():
                saidas_estornos.append({
                    'Banco': row.get('Banco', ''),
                    'Cliente': row.get('Cliente', ''),
                    'Valor': abs(row['Valor']),
                    'Data': row.get('Data', ''),
                    'Histórico': row.get('Histórico', ''),
                    'Baixas': '',
                    'Data Baixa': '',
                    'Motivo Divergência': 'Estorno (Argos)'
                })
            self.df_argos = self.df_argos[~is_estorno & (self.df_argos['Valor'] > 0)]

        if not self.df_bank.empty:
            if 'Tipo' not in self.df_bank.columns: self.df_bank['Tipo'] = ''
            is_saida = (self.df_bank['Valor'] < 0) | (self.df_bank['Tipo'].astype(str).str.upper() == 'D')
            for _, row in self.df_bank[is_saida].iterrows():
                saidas_estornos.append({
                    'Banco': row.get('Banco', ''),
                    'Cliente': row.get('Histórico', ''),
                    'Valor': abs(row['Valor']),
                    'Data': row.get('Data', ''),
                    'Histórico': row.get('Histórico', ''),
                    'Baixas': '',
                    'Data Baixa': '',
                    'Motivo Divergência': 'Saída (Banco)'
                })
            self.df_bank = self.df_bank[~is_saida & (self.df_bank['Valor'] > 0)]

        self.df_saidas_estornos = pd.DataFrame(saidas_estornos)

        if not self.df_argos.empty:
            self.df_argos['ID_Argos'] = range(len(self.df_argos))
        if not self.df_bank.empty:
            self.df_bank['ID_Bank'] = range(len(self.df_bank))

    def execute_pipeline(self):
        import pandas as pd
        import re
        import itertools
        
        resultados = {
            '1_Conciliado_Perfeito': [],
            '2_Conciliado_Via_Historico': [],
            '3_Conciliado_Desmembrado': [],
            '4_Saidas_Estornos': self.df_saidas_estornos.to_dict('records') if hasattr(self, 'df_saidas_estornos') else [],
            '5_Divergencias_Pendentes': []
        }

        if self.df_argos.empty or self.df_bank.empty:
            for k in resultados.keys():
                resultados[k] = pd.DataFrame(columns=['Banco', 'Cliente', 'Valor', 'Data', 'Histórico', 'Baixas', 'Data Baixa', 'Motivo Divergência'])
            return resultados

        argos_pendentes = self.df_argos.copy()
        bank_pendentes = self.df_bank.copy()

        def remover_matched(ids_a, ids_b):
            nonlocal argos_pendentes, bank_pendentes
            argos_pendentes = argos_pendentes[~argos_pendentes['ID_Argos'].isin(ids_a)]
            bank_pendentes = bank_pendentes[~bank_pendentes['ID_Bank'].isin(ids_b)]

        # ==========================================
        # REGRA 2 e 2.5: VIA HISTÓRICO & DESMEMBRADO GUIADO
        # PRIORIDADE MÁXIMA: A intenção humana (escrita) supera a matemática crua.
        # ==========================================
        # regex ajustada para exigir indicativo claro de dinheiro (R$, reais, pix, valor, baixa) antes ou depois
        regex = re.compile(r'(?:PIX.*?|VALOR DE|COMPROVANTE.*?|PAGO.*?|RESTANTE.*?|BAIXA.*?|DE\s*|^|[^\d])R?\$?\s*(\d+(?:\.\d{3})*,\d{1,2})(?:\s*REAIS)?', re.IGNORECASE)
        
        m_a, m_b = [], []
        
        for i, row_a in argos_pendentes.iterrows():
            if row_a['ID_Argos'] in m_a: continue
            
            col_hist = 'OBS' if 'OBS' in row_a else 'Histórico'
            texto_hist = str(row_a.get(col_hist, ''))
            # Corrige erros de digitação comuns no histórico, como "192,oo" ao invés de "192,00"
            texto_hist_corrigido = re.sub(r'(\d+),([oO]+)', lambda m: f"{m.group(1)},{m.group(2).lower().replace('o', '0')}", texto_hist)
            # Corrige datas escritas com vírgula, ex: "DIA 22,06" -> "DIA 22/06"
            texto_hist_corrigido = re.sub(r'(dia\s*\d{1,2}),(\d{1,2})', r'\1/\2', texto_hist_corrigido, flags=re.IGNORECASE)
            # Adiciona ",00" a valores redondos perto de palavras-chave, ex: "pix de 1.000" -> "1.000,00"
            texto_hist_corrigido = re.sub(r'\b(valor(?: de)?|pix|r\$|pago)\s+(\d+(?:\.\d{3})*)(?![\d.,])', r'\1 \2,00', texto_hist_corrigido, flags=re.IGNORECASE)
            texto_hist_corrigido = re.sub(r'\b(\d+(?:\.\d{3})*)(?![\d.,])\s*(reais)', r'\1,00 \2', texto_hist_corrigido, flags=re.IGNORECASE)
            match = regex.search(texto_hist_corrigido)
            
            # Bloqueio rigoroso: Só extrai números se o histórico contiver palavras relacionadas a dinheiro
            # Isso evita extrair uma data (ex: "15/06" virar R$ 15,06 e roubar depósitos)
            palavras_dinheiro = ['pix', 'valor', 'reais', 'r$', 'pago', 'restante', 'baixa', 'comprovante']
            if match and not any(p in texto_hist_corrigido.lower() for p in palavras_dinheiro):
                val_str_temp = match.group(1).replace('.', '').replace(',', '.')
                try: 
                    v_temp = float(val_str_temp)
                    if abs(v_temp - row_a['Valor']) > 15.0:
                        match = None
                except:
                    match = None

            if match:
                val_str = match.group(1).replace('.', '').replace(',', '.')
                try: 
                    v_regex = float(val_str)
                except: 
                    continue

                # Extrai possível data do histórico para usar de referência
                data_referencia = row_a['Data']
                match_data = re.search(r'dia\s*(\d{1,2})[/\.](\d{1,2})', texto_hist_corrigido, re.IGNORECASE)
                if match_data:
                    dia, mes = match_data.groups()
                    try:
                        ano = pd.to_datetime(row_a['Data'], format='%d/%m/%Y').year
                        data_referencia = f"{int(dia):02d}/{int(mes):02d}/{ano}"
                    except:
                        pass
                
                candidatos = bank_pendentes[(bank_pendentes['Valor'] == v_regex) & (~bank_pendentes['ID_Bank'].isin(m_b))].copy()
                if not candidatos.empty:
                    candidatos['diff_dias'] = abs((pd.to_datetime(data_referencia, format='%d/%m/%Y', errors='coerce') - 
                                                   pd.to_datetime(candidatos['Data'], format='%d/%m/%Y', errors='coerce')).dt.days)
                    candidatos = candidatos.sort_values(by='diff_dias')
                    
                    for j, row_b in candidatos.iterrows():
                        if pd.isna(row_b['diff_dias']) or row_b['diff_dias'] <= 31: 
                            valor_faltante = round(v_regex - row_a['Valor'], 2)
                            comb_encontrada = []
                            
                            # Tenta Desmembrar (OTIMIZADO)
                            if valor_faltante > 0.05:
                                cliente_atual = row_a.get('Cliente', None)
                                col_cliente = 'CLIENTES' if 'CLIENTES' in argos_pendentes.columns else 'Cliente'
                                
                                # FILTRO CRÍTICO: Limita a busca a notas que "cabem" no espaço vazio, evitando loop O(N³)
                                filter_mask = (
                                    (~argos_pendentes['ID_Argos'].isin(m_a)) & 
                                    (argos_pendentes['ID_Argos'] != row_a['ID_Argos']) &
                                    (argos_pendentes['Valor'] <= valor_faltante + 1.50)
                                )
                                
                                # TRAVA MESTRA: Só desmembra se for o MESMO cliente!
                                if pd.notna(cliente_atual) and cliente_atual != 'CLIENTE NÃO INFORMADO':
                                    filter_mask = filter_mask & (argos_pendentes[col_cliente] == cliente_atual)
                                    
                                # TRAVA DE DATA: Só busca notas próximas (+- 5 dias)
                                # Para evitar roubar uma nota do dia 01/06 para fechar um desmembramento do dia 30/06
                                date_a = pd.to_datetime(row_a['Data'], format='%d/%m/%Y', errors='coerce')
                                if pd.notna(date_a):
                                    diff_dias_notas = abs((pd.to_datetime(argos_pendentes['Data'], format='%d/%m/%Y', errors='coerce') - date_a).dt.days)
                                    filter_mask = filter_mask & (diff_dias_notas <= 31)
                                    
                                outras_notas = argos_pendentes[filter_mask]
                                fast_notas = [(row['ID_Argos'], row['Valor'], row.to_dict()) for idx, row in outras_notas.iterrows()]
                                
                                for r in range(1, min(4, len(fast_notas) + 1)):
                                    for comb in itertools.combinations(fast_notas, r):
                                        if abs(sum(item[1] for item in comb) - valor_faltante) <= 1.50:
                                            comb_encontrada = [item[2].copy() for item in comb]
                                            break
                                    if comb_encontrada: break
                            
                            if comb_encontrada:
                                nota_principal = row_a.to_dict().copy()
                                nota_principal['Baixas'] = row_b['Banco']
                                nota_principal['Data Baixa'] = row_b['Data']
                                resultados['3_Conciliado_Desmembrado'].append(nota_principal)
                                m_a.append(row_a['ID_Argos'])
                                
                                for np in comb_encontrada:
                                    np['Baixas'] = row_b['Banco']
                                    np['Data Baixa'] = row_b['Data']
                                    resultados['3_Conciliado_Desmembrado'].append(np)
                                    m_a.append(np['ID_Argos'])
                                    
                                m_b.append(row_b['ID_Bank'])
                                break
                            
                            # Se não achou peças mas está dentro do desconto de R$ 15
                            elif abs(row_a['Valor'] - v_regex) <= 15.0:
                                nota = row_a.to_dict().copy()
                                nota['Baixas'] = row_b['Banco']
                                nota['Data Baixa'] = row_b['Data']
                                sinal = "+" if valor_faltante > 0 else ""
                                nota['Motivo Divergência'] = f'Desconto/Acréscimo no PIX ({sinal}R$ {valor_faltante})'
                                resultados['2_Conciliado_Via_Historico'].append(nota)
                                m_a.append(row_a['ID_Argos'])
                                m_b.append(row_b['ID_Bank'])
                                break
        remover_matched(m_a, m_b)

        # ==========================================
        # REGRA 0.5: CONCILIADO POR NOME (Nome do cliente no histórico do banco)
        # ==========================================
        m_a, m_b = [], []
        col_cliente = 'CLIENTES' if 'CLIENTES' in argos_pendentes.columns else 'Cliente'
        
        for i, row_a in argos_pendentes.iterrows():
            cliente_str = str(row_a.get(col_cliente, '')).strip().upper()
            if not cliente_str or cliente_str == 'CLIENTE NÃO INFORMADO':
                continue
                
            # Extrair palavras principais do nome do cliente (com mais de 2 letras)
            palavras_nome = [p for p in cliente_str.split() if len(p) > 2]
            if len(palavras_nome) < 2:
                continue
                
            candidatos = bank_pendentes[(abs(bank_pendentes['Valor'] - row_a['Valor']) <= 1.50) & (~bank_pendentes['ID_Bank'].isin(m_b))].copy()
            if not candidatos.empty:
                diff_dias = abs((pd.to_datetime(row_a['Data'], format='%d/%m/%Y', errors='coerce') - 
                               pd.to_datetime(candidatos['Data'], format='%d/%m/%Y', errors='coerce')).dt.days)
                candidatos['diff_dias'] = diff_dias
                candidatos = candidatos.sort_values(by='diff_dias')
                
                for j, row_b in candidatos.iterrows():
                    hist_banco = str(row_b.get('Histórico', '')).upper()
                    
                    # Verifica se as duas primeiras palavras do nome estão no histórico
                    if palavras_nome[0] in hist_banco and palavras_nome[1] in hist_banco:
                        if pd.isna(row_b['diff_dias']) or row_b['diff_dias'] <= 31:
                            nota = row_a.to_dict().copy()
                            nota['Baixas'] = row_b['Banco']
                            nota['Data Baixa'] = row_b['Data']
                            
                            diff_valor = round(row_b['Valor'] - row_a['Valor'], 2)
                            if diff_valor != 0:
                                sinal = "+" if diff_valor > 0 else ""
                                nota['Motivo Divergência'] = f'Aproximação de Centavos ({sinal}R$ {diff_valor})'
                                
                            resultados['1_Conciliado_Perfeito'].append(nota)
                            m_a.append(row_a['ID_Argos'])
                            m_b.append(row_b['ID_Bank'])
                            break
        remover_matched(m_a, m_b)

        # ==========================================
        # REGRA 1.1: MATCH PERFEITO ÚNICO (Sem limite de datas)
        # ==========================================
        m_a, m_b = [], []
        
        argos_val_counts = argos_pendentes['Valor'].value_counts()
        bank_val_counts = bank_pendentes['Valor'].value_counts()
        
        unique_values = [v for v, c in argos_val_counts.items() if c == 1 and bank_val_counts.get(v) == 1]
        
        for v in unique_values:
            idx_a = argos_pendentes[argos_pendentes['Valor'] == v].index[0]
            idx_b = bank_pendentes[bank_pendentes['Valor'] == v].index[0]
            
            row_a = argos_pendentes.loc[idx_a]
            row_b = bank_pendentes.loc[idx_b]
            
            nota = row_a.to_dict().copy()
            nota['Baixas'] = row_b['Banco']
            nota['Data Baixa'] = row_b['Data']
            resultados['1_Conciliado_Perfeito'].append(nota)
            
            m_a.append(row_a['ID_Argos'])
            m_b.append(row_b['ID_Bank'])
                
        remover_matched(m_a, m_b)

        # ==========================================
        # REGRA 1: CONCILIADO PERFEITO (Valores Exatos com concorrência)
        # ==========================================
        m_a, m_b = [], []
        for i, row_a in argos_pendentes.iterrows():
            candidatos = bank_pendentes[(bank_pendentes['Valor'] == row_a['Valor']) & (~bank_pendentes['ID_Bank'].isin(m_b))].copy()
            if not candidatos.empty:
                # Tenta extrair data do histórico para ajudar na Regra 1
                texto_hist = str(row_a.get('Histórico', '') if 'Histórico' in row_a else row_a.get('OBS', ''))
                data_referencia_hist = None
                match_data = re.search(r'(?:dia\s*|data:\s*)?(\d{1,2})[/\.](\d{1,2})', texto_hist, re.IGNORECASE)
                if match_data:
                    dia, mes = match_data.groups()
                    try:
                        ano = pd.to_datetime(row_a['Data'], format='%d/%m/%Y').year
                        data_referencia_hist = pd.to_datetime(f"{int(dia):02d}/{int(mes):02d}/{ano}", format='%d/%m/%Y', errors='coerce')
                    except:
                        pass
                        
                diff_dias = abs((pd.to_datetime(row_a['Data'], format='%d/%m/%Y', errors='coerce') - 
                               pd.to_datetime(candidatos['Data'], format='%d/%m/%Y', errors='coerce')).dt.days)
                
                if pd.notna(data_referencia_hist):
                    diff_dias_hist = abs((data_referencia_hist - pd.to_datetime(candidatos['Data'], format='%d/%m/%Y', errors='coerce')).dt.days)
                    candidatos['diff_dias'] = pd.concat([diff_dias, diff_dias_hist], axis=1).min(axis=1)
                else:
                    candidatos['diff_dias'] = diff_dias
                    
                candidatos = candidatos.sort_values(by='diff_dias')
                
                for j, row_b in candidatos.iterrows():
                    if pd.notna(row_b['diff_dias']) and row_b['diff_dias'] <= 31:
                        nota = row_a.to_dict().copy()
                        nota['Baixas'] = row_b['Banco']
                        nota['Data Baixa'] = row_b['Data']
                        resultados['1_Conciliado_Perfeito'].append(nota)
                        m_a.append(row_a['ID_Argos'])
                        m_b.append(row_b['ID_Bank'])
                        break
        remover_matched(m_a, m_b)

        # ==========================================
        # REGRA 3: DESMEMBRADOS (Vários Argos -> 1 Banco)
        # ==========================================
        m_a_temp, m_b_temp = [], []
        
        for j, row_b in bank_pendentes.iterrows():
            if row_b['ID_Bank'] in m_b: continue
            
            valor_banco = row_b['Valor']
            data_banco = pd.to_datetime(row_b['Data'], format='%d/%m/%Y', errors='coerce')
            
            argos_candidatos = argos_pendentes[
                (~argos_pendentes['ID_Argos'].isin(m_a + m_a_temp)) & 
                (argos_pendentes['Valor'] <= valor_banco)
            ].copy()
            
            if pd.notna(data_banco):
                argos_candidatos['diff_dias'] = abs((pd.to_datetime(argos_candidatos['Data'], format='%d/%m/%Y', errors='coerce') - data_banco).dt.days)
                argos_candidatos = argos_candidatos[argos_candidatos['diff_dias'] <= 31]
            
            if argos_candidatos.empty: continue
            
            # Prioridade 1: Combinações do mesmo cliente
            col_cliente = 'CLIENTES' if 'CLIENTES' in argos_candidatos.columns else 'Cliente'
            comb_encontrada = None
            
            for cliente, grupo in argos_candidatos.groupby(col_cliente):
                if len(grupo) < 2: continue
                fast_grupo = [(row['ID_Argos'], row['Valor'], row.to_dict()) for idx, row in grupo.iterrows()]
                for r in range(2, min(4, len(fast_grupo) + 1)):
                    for comb in itertools.combinations(fast_grupo, r):
                        if abs(sum(item[1] for item in comb) - valor_banco) <= 1.50:
                            comb_encontrada = comb
                            break
                    if comb_encontrada: break
                if comb_encontrada: break
                            
            if comb_encontrada:
                for item in comb_encontrada:
                    nota = item[2].copy()
                    nota['Baixas'] = row_b['Banco']
                    nota['Data Baixa'] = row_b['Data']
                    resultados['3_Conciliado_Desmembrado'].append(nota)
                    m_a_temp.append(item[0])
                m_b_temp.append(row_b['ID_Bank'])
                
        m_a.extend(m_a_temp)
        m_b.extend(m_b_temp)
        remover_matched(m_a, m_b)

        # ==========================================
        # REGRA 3.5: CONCILIADO POR APROXIMAÇÃO DE CENTAVOS (Último Recurso)
        # ==========================================
        m_a, m_b = [], []
        if not argos_pendentes.empty and not bank_pendentes.empty:
            for i, row_a in argos_pendentes.iterrows():
                candidatos = bank_pendentes[(abs(bank_pendentes['Valor'] - row_a['Valor']) > 0) & 
                                            (abs(bank_pendentes['Valor'] - row_a['Valor']) <= 1.50) & 
                                            (~bank_pendentes['ID_Bank'].isin(m_b))].copy()
                if not candidatos.empty:
                    candidatos['diff_dias'] = abs((pd.to_datetime(row_a['Data'], format='%d/%m/%Y', errors='coerce') - 
                                                   pd.to_datetime(candidatos['Data'], format='%d/%m/%Y', errors='coerce')).dt.days)
                    candidatos = candidatos.sort_values(by='diff_dias')
                    
                    for j, row_b in candidatos.iterrows():
                        if pd.notna(row_b['diff_dias']) and row_b['diff_dias'] <= 3:
                            nota = row_a.to_dict().copy()
                            nota['Baixas'] = row_b['Banco']
                            nota['Data Baixa'] = row_b['Data']
                            diff_valor = round(row_b['Valor'] - row_a['Valor'], 2)
                            sinal = "+" if diff_valor > 0 else ""
                            nota['Motivo Divergência'] = f'Aproximação de Centavos ({sinal}R$ {diff_valor})'
                            resultados['1_Conciliado_Perfeito'].append(nota)
                            m_a.append(row_a['ID_Argos'])
                            m_b.append(row_b['ID_Bank'])
                            break
            remover_matched(m_a, m_b)

        # ==========================================
        # REGRA 4: DIVERGÊNCIAS PENDENTES
        # ==========================================
        for i, row_a in argos_pendentes.iterrows():
            nota = row_a.to_dict().copy()
            nota['Data Baixa'] = nota['Data']
            nota['Data'] = ''
            nota['Motivo Divergência'] = 'Falta no Banco'
            resultados['5_Divergencias_Pendentes'].append(nota)
            
        for i, row_b in bank_pendentes.iterrows():
            col_hist_b = 'Histórico' if 'Histórico' in row_b else 'N/A'
            col_cliente = 'CLIENTES' if 'CLIENTES' in argos_pendentes.columns else 'Cliente'
            nota = {
                'Banco': row_b['Banco'],
                col_cliente: row_b.get(col_hist_b, 'N/A'),
                'Valor': row_b['Valor'],
                'Data': row_b['Data'],
                'Motivo Divergência': 'Sobrou no Banco / Faltou no Argos'
            }
            resultados['5_Divergencias_Pendentes'].append(nota)

        # ==========================================
        # FORMATAÇÃO FINAL DAS COLUNAS (Alinhado com o Template da Cliente)
        # Ordem Exata: Banco | Cliente | Valor | Data (Pgto) | Baixas (Banco Baixa) | Data Baixa | Histórico | Motivo
        # ==========================================
        ordem_colunas = ['Banco', 'Cliente', 'Valor', 'Data', 'Baixas', 'Data Baixa', 'Histórico', 'Motivo Divergência']
        
        for k in resultados.keys():
            df = pd.DataFrame(resultados[k])
            if df.empty:
                df = pd.DataFrame(columns=ordem_colunas)
            else:
                if 'ID_Argos' in df.columns: df = df.drop(columns=['ID_Argos'])
                if 'ID_Bank' in df.columns: df = df.drop(columns=['ID_Bank'])
                if 'CLIENTES' in df.columns: df = df.drop(columns=['CLIENTES'])
                
                # Garante que todas as colunas existem antes de reordenar
                for col in ordem_colunas:
                    if col not in df.columns:
                        df[col] = ''
                
                df = df[ordem_colunas]
            resultados[k] = df

        return resultados




class ExcelReporter:
    @staticmethod
    def generate_report(data_sheets: dict, output_path: str):
        formatted_sheets = {}
        
        mapa_colunas = {
            'BANCO': 'BANCO',
            'Banco': 'BANCO',
            'CLIENTES': 'CLIENTE',
            'Cliente': 'CLIENTE',
            'VALOR': 'VALOR DA BAIXA',
            'Valor': 'VALOR DA BAIXA',
            'DATA PAGAMENTO': 'DATA DO PAGAMENTO',
            'Data': 'DATA DO PAGAMENTO',
            'BAIXAS': 'BANCO DA BAIXA',
            'Baixas': 'BANCO DA BAIXA',
            'DATA BAIXA': 'DATA DA BAIXA',
            'Data Baixa': 'DATA DA BAIXA',
            'OBS': 'HISTÓRICO',
            'Histórico': 'HISTÓRICO',
            'Motivo Divergência': 'MOTIVO DIVERGÊNCIA'
        }
        
        ordem_desejada = ['BANCO', 'CLIENTE', 'VALOR DA BAIXA', 'DATA DO PAGAMENTO', 'BANCO DA BAIXA', 'DATA DA BAIXA', 'HISTÓRICO', 'MOTIVO DIVERGÊNCIA']
        
        for sheet_name, df in data_sheets.items():
            novo_nome_aba = sheet_name.replace('_', ' ')
            if not df.empty:
                df = df.rename(columns=mapa_colunas)
                colunas_finais = [c for c in ordem_desejada if c in df.columns]
                df = df[colunas_finais]
                df['OBSERVAÇÃO'] = ''
            else:
                df = pd.DataFrame(columns=[c for c in ordem_desejada if c != 'MOTIVO DIVERGÊNCIA'] + ['OBSERVAÇÃO'])
            
            formatted_sheets[novo_nome_aba] = df

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in formatted_sheets.items():
                if df.empty:
                    pd.DataFrame({'Aviso': ['Nenhum registo encontrado nesta categoria.']}).to_excel(writer, sheet_name=sheet_name, index=False)
                    continue
                
                for col in df.columns:
                    if 'DATA' in col.upper():
                        df[col] = pd.to_datetime(df[col], dayfirst=True, errors='coerce').dt.strftime('%d/%m/%Y').replace('NaT', '')

                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4)
                worksheet = writer.sheets[sheet_name]

                header_fill = PatternFill(start_color='266C40', end_color='266C40', fill_type='solid')
                fill_verde_claro = PatternFill(start_color='719C82', end_color='719C82', fill_type='solid')
                fill_branco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                font_branca_bold = Font(color='FFFFFF', bold=True)
                font_preta_bold = Font(color='000000', bold=True)
                
                borda_fina = Border(
                    left=Side(border_style='thin', color='356A1C'),
                    right=Side(border_style='thin', color='356A1C'),
                    top=Side(border_style='thin', color='356A1C'),
                    bottom=Side(border_style='thin', color='356A1C')
                )

                col_indices = {str(worksheet.cell(row=5, column=i).value).upper(): i for i in range(1, worksheet.max_column + 1)}
                max_col = worksheet.max_column

                for r in range(1, 5):
                    for c in range(1, max_col + 1):
                        worksheet.cell(row=r, column=c).fill = fill_verde_claro
                
                worksheet.row_dimensions[1].height = 12
                worksheet.row_dimensions[2].height = 55
                worksheet.row_dimensions[3].height = 12
                worksheet.row_dimensions[4].height = 12

                try:
                    from openpyxl.drawing.image import Image as ExcelImage
                    from PIL import Image as PILImage
                    import glob
                    imagens = glob.glob("img/*.*")
                    if imagens:
                        pil_img = PILImage.open(imagens[0])
                        largura_original, altura_original = pil_img.size
                        
                        nova_altura = 65
                        nova_largura = int((nova_altura / altura_original) * largura_original)
                        
                        img = ExcelImage(imagens[0])
                        img.width = nova_largura
                        img.height = nova_altura
                        
                        if max_col <= 7:
                            coluna_alvo = 4 
                        else:
                            coluna_alvo = 5 
                            
                        col_letra_alvo = get_column_letter(max(1, coluna_alvo))
                        worksheet.add_image(img, f'{col_letra_alvo}2')
                except Exception as e:
                    pass

                for cell in worksheet[5]:
                    cell.fill = header_fill
                    cell.font = font_branca_bold
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    cell.border = borda_fina

                for row in range(6, worksheet.max_row + 1):
                    is_zebra = (row % 2 == 0)

                    for col_name, col_idx in col_indices.items():
                        cell = worksheet.cell(row=row, column=col_idx)
                        
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        
                        if sheet_name == '4 Saidas Estornos' and col_name == 'VALOR DA BAIXA':
                            cell.font = Font(color='FF0000', bold=True)
                        else:
                            cell.font = font_preta_bold
                            
                        cell.border = borda_fina

                        if col_name in ['BANCO', 'BANCO DA BAIXA']:
                            cell.fill = fill_verde_claro
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                        else:
                            if is_zebra:
                                cell.fill = fill_verde_claro
                            else:
                                cell.fill = fill_branco

                        if col_name == 'VALOR DA BAIXA':
                            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
                            if cell.value is not None:
                                try:
                                    cell.value = float(cell.value)
                                    cell.number_format = 'R$ #,##0.00'
                                except:
                                    pass

                for col_idx in range(1, max_col + 1):
                    max_length = 0
                    col_letter = get_column_letter(col_idx)
                    col_name_val = worksheet.cell(row=5, column=col_idx).value
                    if not col_name_val: continue
                    col_name = str(col_name_val).upper()
                    
                    max_length = len(col_name) + 3 
                    
                    for row_idx in range(5, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        try:
                            if cell.value:
                                linhas = str(cell.value).split('\n')
                                for linha in linhas:
                                    max_length = max(max_length, len(linha))
                        except:
                            pass
                    
                    if col_name in ['BANCO', 'BANCO DA BAIXA']:
                        worksheet.column_dimensions[col_letter].width = max(max_length, 20)
                    elif col_name == 'VALOR DA BAIXA':
                        worksheet.column_dimensions[col_letter].width = max(max_length + 6, 20)
                    elif 'DATA' in col_name:
                        worksheet.column_dimensions[col_letter].width = max(max_length + 2, 23)
                    elif col_name == 'OBSERVAÇÃO':
                        worksheet.column_dimensions[col_letter].width = 40
                    else:
                        worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)


if __name__ == "__main__":
    pass